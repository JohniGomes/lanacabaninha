import re
import openpyxl

SRC = r"C:\Users\excel\Downloads\ESTOQUE - LA NA CABANINHA.xlsx"
OUT = r"D:\1 - Projetos\La na Cabaninha\lib\estoque-data.ts"

wb = openpyxl.load_workbook(SRC, data_only=True)

items = []


def slug(text):
    text = text.lower()
    text = (
        text.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
        .replace("é", "e").replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o").replace("õ", "o").replace("ô", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text


def add_item(categoria, nome, quantidade, nota=None):
    nome = nome.strip()
    _id = f"est-{slug(categoria)}-{slug(nome)}"
    entry = {"id": _id, "categoria": categoria, "nome": nome.title(), "quantidade": quantidade}
    if nota:
        entry["nota"] = nota
    items.append(entry)


def parse_qty(raw, categoria, nome):
    if isinstance(raw, (int, float)):
        return int(raw), None
    if isinstance(raw, str):
        m = re.search(r"\d+", raw)
        if m:
            return int(m.group()), raw.strip()
    return 0, (str(raw) if raw is not None else None)


def simple_sheet(sheet_name, categoria):
    ws = wb[sheet_name]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    for row in rows[1:]:
        nome, qtd = row[0], row[1]
        if nome is None or str(nome).strip().upper() == "TOTAL":
            continue
        q, nota = parse_qty(qtd, categoria, nome)
        add_item(categoria, str(nome), q, nota)


# CABANAS
simple_sheet("CABANAS", "Cabanas")

# TENDAS - split into Armação / Tecidos
ws = wb["TENDAS"]
rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
current = None
for nome, qtd in rows[1:]:
    if qtd is None:
        current = str(nome).strip().title()
        continue
    categoria = f"Tendas — {current}" if current else "Tendas"
    q, nota = parse_qty(qtd, categoria, nome)
    add_item(categoria, str(nome), q, nota)

# TRAVESSEIROS_FRONHAS - split into Travesseiros / Fronhas
ws = wb["TRAVESSEIROS_FRONHAS"]
rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
# rows: ('TRAVESSEIROS', None), ('87 UNIDADES', None), ('FRONHAS', None), color rows..., ('TOTAL', 277)
add_item("Travesseiros", "Travesseiros (unidades)", 87)
for nome, qtd in rows[3:]:
    if nome is None or str(nome).strip().upper() == "TOTAL":
        continue
    q, nota = parse_qty(qtd, "Fronhas", nome)
    add_item("Fronhas", str(nome), q, nota)

# COBERTORES - two column groups: Casal (A,B) and Solteiro (D,E)
ws = wb["COBERTORES"]
all_rows = list(ws.iter_rows(values_only=True))
for row in all_rows[1:]:
    nome_c, qtd_c = row[0], row[1]
    nome_s, qtd_s = row[3], row[4]
    if nome_c and str(nome_c).strip().upper() != "TOTAL":
        q, nota = parse_qty(qtd_c, "Cobertores Casal", nome_c)
        add_item("Cobertores Casal", str(nome_c), q, nota)
    if nome_s and str(nome_s).strip().upper() != "TOTAL":
        q, nota = parse_qty(qtd_s, "Cobertores Solteiro", nome_s)
        add_item("Cobertores Solteiro", str(nome_s), q, nota)

# Straightforward single-column sheets
simple_sheet("ALMOFADAS", "Almofadas")
simple_sheet("LENÇÓIS", "Lençóis")
simple_sheet("ILUMINAÇÃO", "Iluminação")
simple_sheet("TAPETES", "Tapetes")
simple_sheet("COLCHONETES", "Colchonetes")
simple_sheet("EXTENSÕES", "Extensões")
simple_sheet("BANDEJAS", "Bandejas")
simple_sheet("CAIXAS", "Caixas")

print(f"Total items parsed: {len(items)}")
cats = sorted(set(i["categoria"] for i in items))
print("Categories:", cats)

# Write TS file
lines = []
lines.append("import { EstoqueItem } from \"./types\";")
lines.append("")
lines.append("// Gerado a partir de 'ESTOQUE - LA NA CABANINHA.xlsx' (dados reais da cliente)")
lines.append("export const estoqueInicial: EstoqueItem[] = [")
for it in items:
    nome_js = it["nome"].replace('"', '\\"')
    cat_js = it["categoria"].replace('"', '\\"')
    line = f'  {{ id: "{it["id"]}", categoria: "{cat_js}", nome: "{nome_js}", quantidade: {it["quantidade"]}'
    if "nota" in it and it["nota"]:
        nota_js = str(it["nota"]).replace('"', '\\"')
        line += f', nota: "{nota_js}"'
    line += " },"
    lines.append(line)
lines.append("];")
lines.append("")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"Wrote {OUT}")
