import re
import openpyxl

SRC = r"C:\Users\excel\Downloads\ESTOQUE - LA NA CABANINHA.xlsx"
OUT = r"D:\1 - Projetos\La na Cabaninha\apps-script\estoque-seed.tsv"

wb = openpyxl.load_workbook(SRC, data_only=True)

items = []


def slug(text):
    text = text.lower()
    text = (
        text.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
        .replace("é", "e").replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o").replace("õ", "o").replace("ô", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text


def add_item(categoria, nome, quantidade, nota=None):
    nome = nome.strip()
    _id = f"est-{slug(categoria)}-{slug(nome)}"
    entry = {"id": _id, "categoria": categoria, "nome": nome.title(), "quantidade": quantidade}
    if nota:
        entry["nota"] = nota
    items.append(entry)


def parse_qty(raw, categoria, nome):
    if isinstance(raw, (int, float)):
        return int(raw), None
    if isinstance(raw, str):
        m = re.search(r"\d+", raw)
        if m:
            return int(m.group()), raw.strip()
    return 0, (str(raw) if raw is not None else None)


def simple_sheet(sheet_name, categoria):
    ws = wb[sheet_name]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    for row in rows[1:]:
        nome, qtd = row[0], row[1]
        if nome is None or str(nome).strip().upper() == "TOTAL":
            continue
        q, nota = parse_qty(qtd, categoria, nome)
        add_item(categoria, str(nome), q, nota)


simple_sheet("CABANAS", "Cabanas")

ws = wb["TENDAS"]
rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
current = None
for nome, qtd in rows[1:]:
    if qtd is None:
        current = str(nome).strip().title()
        continue
    categoria = f"Tendas — {current}" if current else "Tendas"
    q, nota = parse_qty(qtd, categoria, nome)
    add_item(categoria, str(nome), q, nota)

ws = wb["TRAVESSEIROS_FRONHAS"]
rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
add_item("Travesseiros", "Travesseiros (unidades)", 87)
for nome, qtd in rows[3:]:
    if nome is None or str(nome).strip().upper() == "TOTAL":
        continue
    q, nota = parse_qty(qtd, "Fronhas", nome)
    add_item("Fronhas", str(nome), q, nota)

ws = wb["COBERTORES"]
all_rows = list(ws.iter_rows(values_only=True))
for row in all_rows[1:]:
    nome_c, qtd_c = row[0], row[1]
    nome_s, qtd_s = row[3], row[4]
    if nome_c and str(nome_c).strip().upper() != "TOTAL":
        q, nota = parse_qty(qtd_c, "Cobertores Casal", nome_c)
        add_item("Cobertores Casal", str(nome_c), q, nota)
    if nome_s and str(nome_s).strip().upper() != "TOTAL":
        q, nota = parse_qty(qtd_s, "Cobertores Solteiro", nome_s)
        add_item("Cobertores Solteiro", str(nome_s), q, nota)

simple_sheet("ALMOFADAS", "Almofadas")
simple_sheet("LENÇÓIS", "Lençóis")
simple_sheet("ILUMINAÇÃO", "Iluminação")
simple_sheet("TAPETES", "Tapetes")
simple_sheet("COLCHONETES", "Colchonetes")
simple_sheet("EXTENSÕES", "Extensões")
simple_sheet("BANDEJAS", "Bandejas")
simple_sheet("CAIXAS", "Caixas")

print(f"Total items parsed: {len(items)}")

lines = ["id\tcategoria\tnome\tquantidade\tnota"]
for it in items:
    nota = it.get("nota", "") or ""
    lines.append(f"{it['id']}\t{it['categoria']}\t{it['nome']}\t{it['quantidade']}\t{nota}")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"Wrote {OUT}")
